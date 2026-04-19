"""
One-time script: add Lab Entry and Lifestyle Entry sheets to TNS_Scan_Data_Master.xlsx
and extend the Unified sheet with matching columns.

Run from the PCA_Pipeline folder:
    python3 _add_lab_lifestyle_sheets.py

This is safe to re-run: if the sheets already exist, they are skipped.
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

XLSX_PATH = Path(__file__).parent.parent / "TNS_Scan_Data_Master.xlsx"
if not XLSX_PATH.exists():
    print(f"ERROR: XLSX not found at {XLSX_PATH}")
    sys.exit(1)

# ── Style helpers ─────────────────────────────────────────────────────────────
HEADER_BLUE  = "0659FF"
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL  = PatternFill("solid", fgColor=HEADER_BLUE)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")   # light yellow = required
OPTIONAL_FILL = PatternFill("solid", fgColor="F2F2F2")   # light grey  = optional
BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)
EXAMPLE_FONT = Font(color="666666", italic=True, size=9)

def _header(ws, col: int, row: int, text: str, required: bool = True):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER
    return cell

def _example(ws, col: int, row: int, val):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = EXAMPLE_FONT
    fill = REQUIRED_FILL if col <= 3 else OPTIONAL_FILL
    cell.fill = fill
    cell.border = BORDER
    return cell

def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width

def _add_dropdown(ws, col: int, start_row: int, end_row: int, formula1: str):
    dv = DataValidation(
        type="list",
        formula1=formula1,
        allow_blank=True,
        showErrorMessage=True,
        error="Choose a value from the list",
        errorTitle="Invalid entry",
    )
    ws.add_data_validation(dv)
    for row in range(start_row, end_row + 1):
        dv.add(ws.cell(row=row, column=col))

def _add_number_validation(ws, col: int, start_row: int, end_row: int,
                            min_val, max_val, allow_blank=True):
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(min_val),
        formula2=str(max_val),
        allow_blank=allow_blank,
        showErrorMessage=True,
        error=f"Enter a number between {min_val} and {max_val}",
        errorTitle="Out of range",
    )
    ws.add_data_validation(dv)
    for row in range(start_row, end_row + 1):
        dv.add(ws.cell(row=row, column=col))


# ── Open workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(str(XLSX_PATH))
existing_sheets = wb.sheetnames
DATA_ROWS = range(3, 203)   # rows 3-202 = 200 data rows

print(f"Opened: {XLSX_PATH.name}")
print(f"Existing sheets: {existing_sheets}")


# ═════════════════════════════════════════════════════════════════════════════
# SHEET: Lab Entry
# ═════════════════════════════════════════════════════════════════════════════
if "Lab Entry" not in existing_sheets:
    ws_lab = wb.create_sheet("Lab Entry")
    ws_lab.sheet_view.showGridLines = True
    ws_lab.freeze_panes = "D3"

    # ── Row 1: section header ─────────────────────────────────────────────
    ws_lab.merge_cells("A1:N1")
    hdr = ws_lab["A1"]
    hdr.value = "TNS Lab Entry — Enter values in the units shown. One row per lab visit."
    hdr.font = Font(bold=True, color="FFFFFF", size=11)
    hdr.fill = PatternFill("solid", fgColor="111127")
    hdr.alignment = Alignment(horizontal="left", vertical="center")
    ws_lab.row_dimensions[1].height = 20

    # ── Row 2: column headers ─────────────────────────────────────────────
    headers = [
        ("A", "client_id",          10), ("B", "scan_date\n(YYYY-MM-DD)", 14),
        ("C", "lab_source",          16), ("D", "total_chol\n(mg/dL)",     10),
        ("E", "HDL\n(mg/dL)",        10), ("F", "LDL\n(mg/dL)",            10),
        ("G", "triglycerides\n(mg/dL)", 13), ("H", "glucose\n(mg/dL, fasting)", 14),
        ("I", "HbA1c\n(%)",          10), ("J", "insulin\n(µIU/mL)",       11),
        ("K", "hs-CRP\n(mg/L)",      10), ("L", "SBP\n(mmHg)",             9),
        ("M", "DBP\n(mmHg)",          9), ("N", "notes",                   25),
    ]
    for i, (_, label, width) in enumerate(headers, 1):
        _header(ws_lab, i, 2, label)
        _set_col_width(ws_lab, i, width)
    ws_lab.row_dimensions[2].height = 35

    # ── Row 3: example row (Jesus Garcia — synthetic labs) ────────────────
    examples = [
        "garcia_jesus", "2026-04-12", "Diagnolab",
        195, 42, 130, 220, 112, 6.1, 12.4, 2.8, 128, 82,
        "Fasting confirmed. Retest in 90 days."
    ]
    for col, val in enumerate(examples, 1):
        _example(ws_lab, col, 3, val)
    ws_lab.row_dimensions[3].height = 16

    # ── Data validations ──────────────────────────────────────────────────
    _add_dropdown(ws_lab, 3, 4, 202,
        '"Diagnolab,CHOPO,Biomédicos de Mérida,Other"')
    _add_number_validation(ws_lab, 4, 4, 202, 50, 500)       # total_chol
    _add_number_validation(ws_lab, 5, 4, 202, 10, 200)        # HDL
    _add_number_validation(ws_lab, 6, 4, 202, 20, 400)        # LDL
    _add_number_validation(ws_lab, 7, 4, 202, 20, 2000)       # triglycerides
    _add_number_validation(ws_lab, 8, 4, 202, 30, 500)        # glucose
    _add_number_validation(ws_lab, 9, 4, 202, 3, 18)          # HbA1c
    _add_number_validation(ws_lab, 10, 4, 202, 0, 300)        # insulin
    _add_number_validation(ws_lab, 11, 4, 202, 0.01, 100)     # hs-CRP
    _add_number_validation(ws_lab, 12, 4, 202, 70, 250)       # SBP
    _add_number_validation(ws_lab, 13, 4, 202, 40, 150)       # DBP

    print("  ✅ Created sheet: Lab Entry")
else:
    print("  SKIP: Lab Entry already exists")


# ═════════════════════════════════════════════════════════════════════════════
# SHEET: Lifestyle Entry
# ═════════════════════════════════════════════════════════════════════════════
if "Lifestyle Entry" not in existing_sheets:
    ws_life = wb.create_sheet("Lifestyle Entry")
    ws_life.sheet_view.showGridLines = True
    ws_life.freeze_panes = "D3"

    # Row 1
    ws_life.merge_cells("A1:K1")
    hdr = ws_life["A1"]
    hdr.value = ("TNS Lifestyle Entry — Collected at intake and each check-in. "
                 "Typical week averages where applicable.")
    hdr.font = Font(bold=True, color="FFFFFF", size=11)
    hdr.fill = PatternFill("solid", fgColor="111127")
    hdr.alignment = Alignment(horizontal="left", vertical="center")
    ws_life.row_dimensions[1].height = 20

    # Row 2: column headers
    lf_headers = [
        ("A", "client_id",            10),
        ("B", "scan_date\n(YYYY-MM-DD)", 14),
        ("C", "vigorous_min\nper_week", 12),
        ("D", "moderate_min\nper_week", 12),
        ("E", "sedentary_hrs\nper_day", 12),
        ("F", "sleep_hrs\nper_night",  12),
        ("G", "smoker",               12),
        ("H", "alcohol_drinks\nper_week", 13),
        ("I", "stress\n(1-10)",       10),
        ("J", "subj_health\n(1-10)",  11),
        ("K", "notes",                25),
    ]
    for i, (_, label, width) in enumerate(lf_headers, 1):
        _header(ws_life, i, 2, label)
        _set_col_width(ws_life, i, width)
    ws_life.row_dimensions[2].height = 35

    # Row 3: example row
    lf_examples = [
        "garcia_jesus", "2026-04-12",
        60, 150, 8.5, 6.5,
        "Never", 5, 7, 5,
        "Desk job. Gym 3x/week."
    ]
    for col, val in enumerate(lf_examples, 1):
        _example(ws_life, col, 3, val)
    ws_life.row_dimensions[3].height = 16

    # Data validations
    _add_number_validation(ws_life, 3, 4, 202, 0, 2000)     # vigorous_min
    _add_number_validation(ws_life, 4, 4, 202, 0, 2000)     # moderate_min
    _add_number_validation(ws_life, 5, 4, 202, 0, 24)       # sed_hours
    _add_number_validation(ws_life, 6, 4, 202, 0, 24)       # sleep_hours
    _add_dropdown(ws_life, 7, 4, 202, '"Never,Former,Current"')
    _add_number_validation(ws_life, 8, 4, 202, 0, 200)      # alcohol
    _add_number_validation(ws_life, 9, 4, 202, 1, 10)       # stress
    _add_number_validation(ws_life, 10, 4, 202, 1, 10)      # subj_health

    print("  ✅ Created sheet: Lifestyle Entry")
else:
    print("  SKIP: Lifestyle Entry already exists")


# ═════════════════════════════════════════════════════════════════════════════
# SHEET: Unified — extend with lab + lifestyle columns
# ═════════════════════════════════════════════════════════════════════════════
UNIFIED_NAME = next(
    (s for s in wb.sheetnames if s.lower().startswith("unified")),
    None
)
if UNIFIED_NAME:
    ws_uni = wb[UNIFIED_NAME]

    # Find the last used column in row 2 (header row)
    max_col = ws_uni.max_column

    # Check if lab columns are already there
    existing_headers = [
        ws_uni.cell(row=2, column=c).value for c in range(1, max_col + 1)
    ]
    has_lab_cols = any("lab_total_chol" in str(h or "") for h in existing_headers)
    has_life_cols = any("lifestyle_vig" in str(h or "") for h in existing_headers)

    # ── Lab columns to append ─────────────────────────────────────────────
    if not has_lab_cols:
        lab_cols = [
            ("lab_total_chol\n(mg/dL)",  "total_chol", 50, 500),
            ("lab_HDL\n(mg/dL)",          "hdl",        10, 200),
            ("lab_LDL\n(mg/dL)",          "ldl",        20, 400),
            ("lab_triglycerides\n(mg/dL)","triglycerides",20, 2000),
            ("lab_glucose\n(mg/dL)",      "glucose",    30, 500),
            ("lab_HbA1c\n(%)",            "hba1c",       3, 18),
            ("lab_insulin\n(µIU/mL)",     "insulin",     0, 300),
            ("lab_hs-CRP\n(mg/L)",        "hscrp",    0.01, 100),
            ("lab_SBP\n(mmHg)",           "sbp",        70, 250),
            ("lab_DBP\n(mmHg)",           "dbp",        40, 150),
        ]
        # Need to figure out which column index client_id is to build SUMPRODUCT
        # We'll assume col A = scan_date, col B = client_id based on typical Unified layout
        # Use a simpler approach: IFERROR VLOOKUP-style formula placeholder
        # The sheet has client_id in a column; user can adjust if layout differs
        for i, (header, field, mn, mx) in enumerate(lab_cols):
            c = max_col + 1 + i
            ws_uni.cell(row=2, column=c, value=header).font = HEADER_FONT
            ws_uni.cell(row=2, column=c).fill = PatternFill("solid", fgColor="1F4E79")
            ws_uni.cell(row=2, column=c).alignment = HEADER_ALIGN
            ws_uni.cell(row=2, column=c).border = BORDER
            ws_uni.column_dimensions[get_column_letter(c)].width = 12

            # Row 1: section label on first lab column
            if i == 0:
                ws_uni.cell(row=1, column=c, value="▼ Lab Values (auto-merged from Lab Entry)")
                ws_uni.cell(row=1, column=c).font = Font(bold=True, color="CDDEFF", size=9)
                ws_uni.cell(row=1, column=c).fill = PatternFill("solid", fgColor="0D2B4E")

            # Rows 3+: IFERROR formula — pulls from Lab Entry matching client_id + date ±3 days
            # Assumes Unified col A = scan_date (date), col B = client_id (text)
            # Lab Entry: col A = client_id, col B = scan_date, col D+ = lab values
            lab_col_offset = 4 + [
                "total_chol","hdl","ldl","triglycerides",
                "glucose","hba1c","insulin","hscrp","sbp","dbp"
            ].index(field)
            lab_col_letter = get_column_letter(lab_col_offset)

            for row in range(3, 203):
                formula = (
                    f"=IFERROR(INDEX('Lab Entry'!{lab_col_letter}:{lab_col_letter},"
                    f"SUMPRODUCT(('Lab Entry'!A$1:A$202=$B{row})"
                    f"*(ABS('Lab Entry'!B$1:B$202-$A{row})<=3)"
                    f"*ROW('Lab Entry'!A$1:A$202))),"
                    f'"")'
                )
                ws_uni.cell(row=row, column=c, value=formula)

        max_col += len(lab_cols)
        print(f"  ✅ Added {len(lab_cols)} lab columns to Unified")

    # ── Lifestyle columns to append ───────────────────────────────────────
    if not has_life_cols:
        life_cols = [
            ("lifestyle_vig_min\nper_week",  "vigorous_min_per_week",  3),
            ("lifestyle_mod_min\nper_week",  "moderate_min_per_week",  4),
            ("lifestyle_sed_hrs\nper_day",   "sedentary_hrs_per_day",  5),
            ("lifestyle_sleep_hrs",          "sleep_hrs_per_night",    6),
            ("lifestyle_smoker",             "smoker",                 7),
            ("lifestyle_alcohol\ndrinks/wk", "alcohol_drinks_per_week",8),
            ("lifestyle_stress\n(1-10)",     "stress_1to10",           9),
            ("lifestyle_subj_health\n(1-10)","subj_health_1to10",     10),
            ("data_completeness\n(scan/lab/life)", "_completeness",    None),
        ]
        for i, (header, field, life_col_num) in enumerate(life_cols):
            c = max_col + 1 + i
            ws_uni.cell(row=2, column=c, value=header).font = HEADER_FONT
            ws_uni.cell(row=2, column=c).fill = PatternFill("solid", fgColor="1A4731")
            ws_uni.cell(row=2, column=c).alignment = HEADER_ALIGN
            ws_uni.cell(row=2, column=c).border = BORDER
            ws_uni.column_dimensions[get_column_letter(c)].width = 13

            if i == 0:
                ws_uni.cell(row=1, column=c,
                            value="▼ Lifestyle (auto-merged from Lifestyle Entry)")
                ws_uni.cell(row=1, column=c).font = Font(bold=True, color="CDDEFF", size=9)
                ws_uni.cell(row=1, column=c).fill = PatternFill("solid", fgColor="0D2B1A")

            if field == "_completeness":
                # Text summary formula: "scan/lab/life" present
                for row in range(3, 203):
                    ws_uni.cell(row=row, column=c, value=f"=IF($B{row}=\"\",\"\",\"scan\")")
            elif life_col_num is not None:
                life_col_letter = get_column_letter(life_col_num)
                for row in range(3, 203):
                    formula = (
                        f"=IFERROR(INDEX('Lifestyle Entry'!{life_col_letter}:{life_col_letter},"
                        f"SUMPRODUCT(('Lifestyle Entry'!A$1:A$202=$B{row})"
                        f"*(ABS('Lifestyle Entry'!B$1:B$202-$A{row})<=3)"
                        f"*ROW('Lifestyle Entry'!A$1:A$202))),"
                        f'"")'
                    )
                    ws_uni.cell(row=row, column=c, value=formula)

        print(f"  ✅ Added {len(life_cols)} lifestyle columns to Unified")

    print(f"  Unified sheet '{UNIFIED_NAME}' updated")
else:
    print("  NOTE: Unified sheet not found — lab/lifestyle columns not added to Unified")


# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(str(XLSX_PATH))
print(f"\n✅ Saved: {XLSX_PATH.name}")
print(f"   Sheets: {wb.sheetnames}")
